from __future__ import annotations

"""FastAPI web layer for monitoring, journaling and trade analysis.

新人接手建议先从本文件读起：
1. `create_app` 负责组装整个 Web 应用
2. `lifespan` 负责启动时初始化数据库和后台轮询
3. `/dashboard`、`/trades`、`/history` 分别对应三个核心页面
4. 多账号逻辑通过 `_get_authenticated_user` 和数据库里的 `owner_username` 实现数据隔离
"""

import asyncio
from contextlib import asynccontextmanager
from datetime import UTC, date, datetime, timedelta
import html
from io import BytesIO
import json
from pathlib import Path
from zoneinfo import ZoneInfo
import secrets
import sqlite3
from urllib.parse import quote

from fastapi import FastAPI, Form, Request
from fastapi.responses import HTMLResponse, RedirectResponse, Response, StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
import uvicorn

from .config import AppSettings, load_settings
from .database import (
    add_monitored_stock,
    add_trade_analysis,
    add_trade_record,
    clear_login_guard_state,
    clear_login_guard_states_for_username,
    consume_login_unlock_code,
    create_user,
    get_admin_user,
    get_alert_history,
    get_email_settings,
    get_latest_trade_analysis,
    get_login_guard_state,
    get_portfolio_settings,
    get_login_unlock_code,
    get_quant_settings,
    get_snapshot,
    get_snapshots,
    get_unread_alerts,
    get_user,
    initialize_database,
    list_recent_login_events,
    list_trade_records,
    list_trade_records_for_symbol,
    list_trade_records_paginated,
    list_users,
    mark_alert_as_read,
    record_failed_login,
    record_login_event,
    remove_monitored_stock,
    save_portfolio_settings,
    save_email_settings,
    save_login_unlock_code,
    save_quant_settings,
    update_user_password,
    delete_trade_record,
    delete_trade_records_for_symbol,
)
from .mailer import build_login_unlock_email_body, build_test_email_body, build_trade_analysis_email_body, send_message
from .monitor import StockMonitor, parse_stock_symbols
from .quant import available_quant_models, normalize_selected_models, normalize_strategy_params
from .security import hash_password, password_hash_needs_rehash, verify_password
from .trade_advisor import TradeAdvisor, build_portfolio_profile, build_position_summary, build_stock_comprehensive_advice


BASE_DIR = Path(__file__).resolve().parent
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))


SESSION_IDLE_TIMEOUT_SECONDS = 30 * 60


def _to_local_datetime(raw_value: str | None, timezone_name: str = "Asia/Shanghai") -> datetime | None:
    if not raw_value:
        return None
    try:
        parsed = datetime.fromisoformat(raw_value)
    except ValueError:
        return None
    target_zone = ZoneInfo(timezone_name)
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=target_zone)
    return parsed.astimezone(target_zone)


def _format_snapshot_timestamp(raw_value: str | None, timezone_name: str = "Asia/Shanghai") -> str:
    """Show a compact timestamp in the dashboard table using local server timezone."""

    parsed = _to_local_datetime(raw_value, timezone_name)
    if parsed is None:
        return raw_value[:16].replace("T", " ") if raw_value else "-"
    now = datetime.now(ZoneInfo(timezone_name))
    if parsed.date() == now.date():
        return parsed.strftime("%H:%M")
    return parsed.strftime("%m-%d %H:%M")


def _format_full_timestamp(raw_value: str | None, timezone_name: str = "Asia/Shanghai") -> str:
    parsed = _to_local_datetime(raw_value, timezone_name)
    if parsed is None:
        return raw_value[:19].replace("T", " ") if raw_value else "-"
    return parsed.strftime("%Y-%m-%d %H:%M:%S")


def _format_stock_display_name(name: str | None, chunk_size: int = 4) -> str:
    """Insert soft-wrap opportunities only for long pure-Chinese stock names."""

    if not name:
        return "-"
    escaped = html.escape(name)
    if len(name) <= chunk_size:
        return escaped
    if not all("一" <= char <= "鿿" for char in name):
        return escaped
    return "<wbr>".join(html.escape(name[index : index + chunk_size]) for index in range(0, len(name), chunk_size))


templates.env.globals["format_snapshot_timestamp"] = _format_snapshot_timestamp
templates.env.globals["format_full_timestamp"] = _format_full_timestamp
templates.env.globals["format_stock_display_name"] = _format_stock_display_name


def _decorate_snapshot_for_display(snapshot: object) -> dict[str, object]:
    payload = dict(snapshot)
    payload.update(build_stock_comprehensive_advice(payload))
    return payload


def create_app(settings: AppSettings | None = None) -> FastAPI:
    resolved_settings = settings or load_settings()

    @asynccontextmanager
    async def lifespan(app: FastAPI):
        initialize_database(resolved_settings)
        app.state.monitor_task = asyncio.create_task(_monitor_loop(app))
        try:
            yield
        finally:
            task = app.state.monitor_task
            if task is not None:
                task.cancel()
                try:
                    await task
                except asyncio.CancelledError:
                    pass

    app = FastAPI(title=resolved_settings.app_name, lifespan=lifespan)
    app.state.settings = resolved_settings
    app.state.monitor = StockMonitor(resolved_settings)
    app.state.trade_advisor = TradeAdvisor(resolved_settings)
    app.state.monitor_task = None

    @app.middleware("http")
    async def refresh_authenticated_session(request: Request, call_next):
        response = await call_next(request)
        if request.url.path in {"/login", "/login/unlock/request", "/login/unlock/confirm", "/healthz"}:
            return response
        current_user = _get_authenticated_user(request, resolved_settings)
        if current_user is None:
            return response
        response.set_cookie(
            resolved_settings.session_cookie_name,
            _build_session_cookie_value(current_user),
            httponly=True,
            samesite="lax",
        )
        return response

    @app.get("/", response_class=HTMLResponse)
    async def index(request: Request) -> Response:
        if _get_authenticated_user(request, resolved_settings) is not None:
            return RedirectResponse(url="/dashboard", status_code=303)
        return RedirectResponse(url="/login", status_code=303)

    @app.get("/login", response_class=HTMLResponse)
    async def login_page(
        request: Request,
        error: str | None = None,
        message: str | None = None,
        unlock_username: str | None = None,
    ) -> Response:
        return _render_login_page(
            request,
            error=error,
            message=message,
            unlock_username=unlock_username or "",
        )

    @app.post("/login")
    async def login(
        request: Request,
        username: str = Form(...),
        password: str = Form(...),
    ) -> Response:
        normalized_username = username.strip()
        client_subject = _login_subject(request, normalized_username)
        guard_state = get_login_guard_state(resolved_settings.db_path, client_subject)
        if guard_state and guard_state["locked_until"]:
            locked_until = datetime.fromisoformat(guard_state["locked_until"])
            now = datetime.now(locked_until.tzinfo)
            if locked_until > now:
                remaining_seconds = max(1, int((locked_until - now).total_seconds()))
                remaining_minutes = (remaining_seconds + 59) // 60
                return _render_login_page(
                    request,
                    error=(
                        f"登录失败次数过多，当前访问已被临时锁定，请约 {remaining_minutes} 分钟后再试。"
                        " 如已绑定邮箱，可直接在下方发送验证码解封。"
                    ),
                    unlock_username=normalized_username,
                    status_code=429,
                )

        user = get_user(resolved_settings.db_path, normalized_username)
        if user is None or not verify_password(password, user["password_hash"]):
            record_failed_login(resolved_settings.db_path, client_subject)
            return _render_login_page(
                request,
                error="用户名或密码错误",
                unlock_username=normalized_username,
                status_code=400,
            )

        clear_login_guard_state(resolved_settings.db_path, client_subject)
        if password_hash_needs_rehash(user["password_hash"]):
            update_user_password(resolved_settings.db_path, user["username"], hash_password(password))
            user = get_user(resolved_settings.db_path, normalized_username)
        if user is None:
            return _render_login_page(
                request,
                error="登录态初始化失败，请重试",
                unlock_username=normalized_username,
                status_code=500,
            )

        record_login_event(
            resolved_settings.db_path,
            user["username"],
            _request_client_host(request),
            _request_user_agent(request),
        )
        response = RedirectResponse(url="/dashboard", status_code=303)
        response.set_cookie(
            resolved_settings.session_cookie_name,
            _build_session_cookie_value(user),
            httponly=True,
            samesite="lax",
        )
        return response

    @app.post("/login/unlock/request")
    async def request_login_unlock(
        request: Request,
        username: str = Form(...),
    ) -> Response:
        normalized_username = username.strip()
        user = get_user(resolved_settings.db_path, normalized_username)
        if user is None:
            return _render_login_page(
                request,
                error="未找到该账号，无法发送解封验证码。",
                unlock_username=normalized_username,
                status_code=404,
            )

        email_settings = get_email_settings(resolved_settings.db_path, normalized_username)
        target_email = _resolve_unlock_email_target(email_settings)
        if not target_email:
            return _render_login_page(
                request,
                error="该账号尚未配置可用邮箱，暂时无法自助解封，请联系管理员。",
                unlock_username=normalized_username,
                status_code=400,
            )

        verification_code = f"{secrets.randbelow(1_000_000):06d}"
        expires_minutes = 10
        expires_at = (datetime.now(UTC) + timedelta(minutes=expires_minutes)).isoformat()
        save_login_unlock_code(
            resolved_settings.db_path,
            normalized_username,
            verification_code,
            expires_at,
        )
        unlock_email_settings = dict(email_settings)
        unlock_email_settings["recipient_email"] = target_email
        result = send_message(
            unlock_email_settings,
            subject=f"[登录解封] {normalized_username} 验证码",
            body=build_login_unlock_email_body(normalized_username, verification_code, expires_minutes),
        )
        if not result.success:
            return _render_login_page(
                request,
                error=f"解封邮件发送失败：{result.error}",
                unlock_username=normalized_username,
                status_code=400,
            )
        masked_email = _mask_email_address(target_email)
        return _render_login_page(
            request,
            message=f"验证码已发送到 {masked_email}，请输入验证码完成解封。",
            unlock_username=normalized_username,
        )

    @app.post("/login/unlock/confirm")
    async def confirm_login_unlock(
        request: Request,
        username: str = Form(...),
        verification_code: str = Form(...),
    ) -> Response:
        normalized_username = username.strip()
        unlock_row = get_login_unlock_code(resolved_settings.db_path, normalized_username)
        if unlock_row is None:
            return _render_login_page(
                request,
                error="请先发送解封验证码。",
                unlock_username=normalized_username,
                status_code=400,
            )
        if unlock_row["consumed_at"]:
            return _render_login_page(
                request,
                error="该验证码已使用，请重新发送新的验证码。",
                unlock_username=normalized_username,
                status_code=400,
            )
        expires_at = datetime.fromisoformat(unlock_row["expires_at"])
        if expires_at <= datetime.now(expires_at.tzinfo):
            return _render_login_page(
                request,
                error="验证码已过期，请重新发送新的验证码。",
                unlock_username=normalized_username,
                status_code=400,
            )
        if verification_code.strip() != str(unlock_row["verification_code"]):
            return _render_login_page(
                request,
                error="验证码不正确，请重新输入。",
                unlock_username=normalized_username,
                status_code=400,
            )

        clear_login_guard_states_for_username(resolved_settings.db_path, normalized_username)
        consume_login_unlock_code(resolved_settings.db_path, normalized_username)
        return _render_login_page(
            request,
            message="邮箱验证成功，账号锁定已解除，请重新登录。",
            unlock_username=normalized_username,
        )

    @app.get("/logout")
    async def logout(reason: str | None = None) -> RedirectResponse:
        target_url = "/login"
        if reason == "idle":
            target_url = "/login?message=%E7%94%B1%E4%BA%8E30%E5%88%86%E9%92%9F%E6%97%A0%E6%93%8D%E4%BD%9C%EF%BC%8C%E5%B7%B2%E8%87%AA%E5%8A%A8%E9%80%80%E5%87%BA%E7%99%BB%E5%BD%95%E3%80%82"
        response = RedirectResponse(url=target_url, status_code=303)
        response.delete_cookie(resolved_settings.session_cookie_name)
        return response

    @app.get("/trades", response_class=HTMLResponse)
    async def trades_page(
        request: Request,
        symbol: str | None = None,
        page: int = 1,
        message: str | None = None,
        message_type: str = "info",
    ) -> Response:
        current_user = _require_login(request, resolved_settings)
        if isinstance(current_user, RedirectResponse):
            return current_user

        selected_symbol = (symbol or "").strip()
        owner_username = current_user["username"]
        portfolio_settings = get_portfolio_settings(resolved_settings.db_path, owner_username)
        trade_records, total_trade_records = list_trade_records_paginated(
            resolved_settings.db_path,
            owner_username,
            selected_symbol or None,
            page=page,
            page_size=10,
        )
        all_trade_records = list_trade_records(resolved_settings.db_path, owner_username, None)
        snapshots = get_snapshots(resolved_settings.db_path, owner_username)
        latest_analysis_row = get_latest_trade_analysis(
            resolved_settings.db_path,
            owner_username,
            selected_symbol or None,
        )
        latest_analysis = None
        position_summary = None
        if latest_analysis_row is not None:
            latest_analysis = _serialize_latest_analysis_row(latest_analysis_row)
            position_summary = json.loads(latest_analysis_row["position_summary"])
        elif selected_symbol:
            position_summary = build_position_summary(
                list_trade_records_for_symbol(resolved_settings.db_path, owner_username, selected_symbol)
            )

        portfolio_profile = build_portfolio_profile(all_trade_records, snapshots, float(portfolio_settings["total_investment_amount"] or 0.0))

        return templates.TemplateResponse(
            request,
            "trades.html",
            _base_template_context(
                request,
                current_user,
                message=message,
                message_type=message_type,
                page_title="交易复盘",
                selected_symbol=selected_symbol,
                snapshots=snapshots,
                trade_records=trade_records,
                latest_analysis=latest_analysis,
                position_summary=position_summary,
                portfolio_profile=portfolio_profile,
                portfolio_settings=portfolio_settings,
                trade_pagination={
                    "page": max(page, 1),
                    "page_size": 10,
                    "total": total_trade_records,
                    "total_pages": max((total_trade_records + 9) // 10, 1),
                },
            ),
        )

    @app.post("/settings/portfolio")
    async def save_portfolio_setting(
        request: Request,
        total_investment_amount: float = Form(0.0),
        next_path: str = Form("/trades"),
    ) -> RedirectResponse:
        current_user = _require_login(request, resolved_settings)
        if isinstance(current_user, RedirectResponse):
            return current_user
        save_portfolio_settings(
            resolved_settings.db_path,
            current_user["username"],
            max(float(total_investment_amount), 0.0),
        )
        return _redirect_with_message(next_path or "/trades", "股市总投入金额已更新")

    @app.post("/trades/records/{trade_id}/delete")
    async def delete_trade_record_route(request: Request, trade_id: int) -> RedirectResponse:
        current_user = _require_login(request, resolved_settings)
        if isinstance(current_user, RedirectResponse):
            return current_user
        symbol = (request.query_params.get("symbol") or "").strip()
        page_value = request.query_params.get("page") or "1"
        delete_trade_record(resolved_settings.db_path, current_user["username"], trade_id)
        redirect_url = f"/trades?page={quote(page_value)}"
        if symbol:
            redirect_url = f"/trades?symbol={quote(symbol)}&page={quote(page_value)}"
        return _redirect_with_message(redirect_url, "交易记录已删除")

    @app.post("/trades/positions/{symbol}/delete")
    async def delete_position_records_route(request: Request, symbol: str) -> RedirectResponse:
        current_user = _require_login(request, resolved_settings)
        if isinstance(current_user, RedirectResponse):
            return current_user
        delete_trade_records_for_symbol(resolved_settings.db_path, current_user["username"], symbol)
        return _redirect_with_message("/trades", f"{symbol} 持仓记录已清空")

    @app.get("/healthz")
    async def healthz() -> dict[str, str]:
        return {"status": "ok"}

    @app.get("/trades/export")
    async def export_trades(request: Request, symbol: str | None = None) -> Response:
        current_user = _require_login(request, resolved_settings)
        if isinstance(current_user, RedirectResponse):
            return current_user

        owner_username = current_user["username"]
        selected_symbol = (symbol or "").strip() or None
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "交易流水"
        worksheet.append(["成交时间", "股票代码", "方向", "价格", "数量", "备注", "录入时间"])
        trade_records = list_trade_records(resolved_settings.db_path, owner_username, selected_symbol)
        for trade in trade_records:
            worksheet.append(
                [
                    trade["traded_at"],
                    trade["symbol"],
                    "买入" if trade["side"] == "buy" else "卖出",
                    float(trade["price"]),
                    int(trade["quantity"]),
                    trade["note"] or "",
                    trade["created_at"],
                ]
            )

        portfolio_profile = build_portfolio_profile(
            list_trade_records(resolved_settings.db_path, owner_username, None),
            get_snapshots(resolved_settings.db_path, owner_username),
        )
        latest_analysis_row = get_latest_trade_analysis(resolved_settings.db_path, owner_username, selected_symbol)
        if latest_analysis_row is not None:
            analysis_sheet = workbook.create_sheet(title="最新分析")
            analysis = json.loads(latest_analysis_row["analysis_json"])
            market_snapshot = _decorate_snapshot_for_display(json.loads(latest_analysis_row["market_snapshot"]))
            analysis_sheet.append(["股票代码", latest_analysis_row["symbol"]])
            analysis_sheet.append(["分析来源", latest_analysis_row["analysis_provider"]])
            analysis_sheet.append(["模型", latest_analysis_row["model_name"]])
            analysis_sheet.append(["状态", latest_analysis_row["status"]])
            analysis_sheet.append(["生成时间", latest_analysis_row["created_at"]])
            analysis_sheet.append(["结论", analysis["summary"]])
            analysis_sheet.append(["合理性判断", analysis["judgment"]])
            analysis_sheet.append(["仓位建议", analysis["position_advice"]])
            analysis_sheet.append(["推荐买入价", analysis.get("recommended_buy_price_range", "")])
            analysis_sheet.append(["推荐买入等级", f"{analysis.get('buy_recommendation_level', '')}/10 {analysis.get('buy_recommendation_level_label', '')}".strip()])
            analysis_sheet.append(["推荐卖出价", analysis.get("recommended_sell_price_range", "")])
            analysis_sheet.append(["推荐卖出等级", f"{analysis.get('sell_recommendation_level', '')}/10 {analysis.get('sell_recommendation_level_label', '')}".strip()])
            analysis_sheet.append(["观望关注价", analysis.get("watch_price_range", "")])
            analysis_sheet.append(["DCF代理内在价值", market_snapshot.get("dcf_intrinsic_value") or ""])
            analysis_sheet.append(["DCF估值偏差(%)", market_snapshot.get("dcf_valuation_gap_pct") or ""])
            analysis_sheet.append(["组合建议仓位", portfolio_profile.get("recommended_holding_ratio", "")])
            analysis_sheet.append(["组合调仓目标", portfolio_profile.get("target_holding_ratio_mid", "")])
            analysis_sheet.append(["组合综合建议", portfolio_profile.get("comprehensive_advice", "")])
            analysis_sheet.append(["优先减仓对象", " | ".join(portfolio_profile.get("priority_reduce_positions", []))])
            analysis_sheet.append(["优先加仓对象", " | ".join(portfolio_profile.get("priority_add_positions", []))])
            analysis_sheet.append(["置信度", analysis["confidence"]])

        portfolio_sheet = workbook.create_sheet(title="组合总览")
        portfolio_sheet.append(["当前持仓比例", portfolio_profile.get("holding_ratio", 0)])
        portfolio_sheet.append(["模型建议仓位", portfolio_profile.get("recommended_holding_ratio", "")])
        portfolio_sheet.append(["目标仓位中枢", portfolio_profile.get("target_holding_ratio_mid", "")])
        portfolio_sheet.append(["持仓风格", portfolio_profile.get("holding_style", "")])
        portfolio_sheet.append(["风险程度", portfolio_profile.get("risk_level", "")])
        portfolio_sheet.append(["组合建议", portfolio_profile.get("comprehensive_advice", "")])
        portfolio_sheet.append(["调仓建议", " | ".join(portfolio_profile.get("overall_adjustment_suggestions", []))])
        portfolio_sheet.append(["优先减仓", " | ".join(portfolio_profile.get("priority_reduce_positions", []))])
        portfolio_sheet.append(["优先加仓", " | ".join(portfolio_profile.get("priority_add_positions", []))])
        portfolio_sheet.append([])
        portfolio_sheet.append(["股票", "仓位占比", "建议加仓价", "建议减仓价", "建议止损价", "DCF代理内在价值", "DCF估值偏差(%)", "风险", "建议"])
        for item in portfolio_profile.get("active_positions", []):
            portfolio_sheet.append([
                item["symbol"],
                item.get("weight_pct", 0.0),
                item.get("suggested_add_price") or "",
                item.get("suggested_reduce_price") or "",
                item.get("suggested_stop_loss_price") or "",
                item.get("dcf_intrinsic_value") or "",
                item.get("dcf_valuation_gap_pct") or "",
                item.get("risk_level", ""),
                item.get("action", ""),
            ])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        filename_symbol = selected_symbol or "all"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": (
                    f'attachment; filename="trade_records_{owner_username}_{filename_symbol}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
                )
            },
        )

    @app.get("/dashboard", response_class=HTMLResponse)
    async def dashboard(
        request: Request,
        message: str | None = None,
        message_type: str = "info",
    ) -> Response:
        current_user = _require_login(request, resolved_settings)
        if isinstance(current_user, RedirectResponse):
            return current_user

        monitor: StockMonitor = app.state.monitor
        owner_username = current_user["username"]
        status = monitor.get_market_status()
        raw_snapshots = await _load_dashboard_snapshots(app, owner_username)
        snapshots = [_decorate_snapshot_for_display(item) for item in raw_snapshots]
        portfolio_settings = get_portfolio_settings(resolved_settings.db_path, owner_username)
        portfolio_profile = build_portfolio_profile(
            list_trade_records(resolved_settings.db_path, owner_username, None),
            raw_snapshots,
            float(portfolio_settings["total_investment_amount"] or 0.0),
        )
        email_settings = get_email_settings(resolved_settings.db_path, owner_username)
        unread_alerts = get_unread_alerts(resolved_settings.db_path, owner_username) if status.is_market_open else []
        return templates.TemplateResponse(
            request,
            "dashboard.html",
            _base_template_context(
                request,
                current_user,
                message=message,
                message_type=message_type,
                page_title="监控面板",
                portfolio_profile=portfolio_profile,
                monitor=monitor,
                price_column_label="最新价" if status.is_market_open else "最近收盘/最新可用价",
                snapshots=snapshots,
                status=status,
                system_status=_build_system_status(monitor, email_settings),
                unread_alerts=unread_alerts,
            ),
        )

    @app.get("/settings", response_class=HTMLResponse)
    async def settings_page(
        request: Request,
        message: str | None = None,
        message_type: str = "info",
    ) -> Response:
        current_user = _require_login(request, resolved_settings)
        if isinstance(current_user, RedirectResponse):
            return current_user

        owner_username = current_user["username"]
        email_settings = get_email_settings(resolved_settings.db_path, owner_username)
        quant_settings = get_quant_settings(resolved_settings.db_path, owner_username)
        quant_strategy_params = normalize_strategy_params(json.loads(quant_settings["strategy_params"] or "{}"))
        selected_models = normalize_selected_models(json.loads(quant_settings["selected_models"] or "[]"))
        recent_login_events = list_recent_login_events(resolved_settings.db_path, owner_username)
        latest_login_event = recent_login_events[0] if recent_login_events else None
        previous_login_event = recent_login_events[1] if len(recent_login_events) > 1 else None
        return templates.TemplateResponse(
            request,
            "settings.html",
            _base_template_context(
                request,
                current_user,
                message=message,
                message_type=message_type,
                page_title="设置",
                email_settings=email_settings,
                quant_settings=quant_settings,
                quant_model_options=available_quant_models(),
                quant_selected_models=selected_models,
                quant_strategy_params=quant_strategy_params,
                recent_login_events=recent_login_events,
                latest_login_event=latest_login_event,
                previous_login_event=previous_login_event,
                users=list_users(resolved_settings.db_path) if current_user["is_admin"] else [],
            ),
        )

    @app.post("/dashboard/refresh")
    async def refresh_dashboard(request: Request) -> RedirectResponse:
        current_user = _require_login(request, resolved_settings)
        if isinstance(current_user, RedirectResponse):
            return current_user

        snapshots = get_snapshots(resolved_settings.db_path, current_user["username"])
        if not snapshots:
            return _redirect_with_message("/dashboard", "当前账号还没有监控股票")

        refresh_failed_symbols: list[str] = []
        for item in snapshots:
            try:
                await asyncio.to_thread(app.state.monitor.refresh_symbol_snapshot, current_user["username"], item["symbol"])
            except Exception:
                refresh_failed_symbols.append(item["symbol"])
        if refresh_failed_symbols:
            return _redirect_with_message(
                "/dashboard",
                f"已完成部分刷新，失败股票：{', '.join(refresh_failed_symbols)}",
            )
        return _redirect_with_message("/dashboard", f"已手动刷新 {len(snapshots)} 只股票")

    @app.post("/stocks")
    async def add_stocks(request: Request, symbols_text: str = Form(...)) -> RedirectResponse:
        current_user = _require_login(request, resolved_settings)
        if isinstance(current_user, RedirectResponse):
            return current_user

        owner_username = current_user["username"]
        valid_symbols, invalid_symbols = parse_stock_symbols(symbols_text)
        if not valid_symbols:
            return _redirect_with_message("/settings", "请输入6位有效A股代码")
        refresh_failed_symbols: list[str] = []
        for symbol in valid_symbols:
            add_monitored_stock(resolved_settings.db_path, owner_username, symbol)
            try:
                await asyncio.to_thread(app.state.monitor.refresh_symbol_snapshot, owner_username, symbol)
            except Exception:
                refresh_failed_symbols.append(symbol)
        if invalid_symbols:
            return _redirect_with_message(
                "/settings",
                f"已加入 {len(valid_symbols)} 只股票；以下代码无效：{', '.join(invalid_symbols)}",
            )
        if refresh_failed_symbols:
            return _redirect_with_message(
                "/settings",
                f"已加入 {len(valid_symbols)} 只股票；以下股票暂未拉到最新数据：{', '.join(refresh_failed_symbols)}",
            )
        return _redirect_with_message("/settings", f"已加入 {len(valid_symbols)} 只股票")

    @app.post("/stocks/{symbol}/delete")
    async def delete_stock(request: Request, symbol: str) -> RedirectResponse:
        current_user = _require_login(request, resolved_settings)
        if isinstance(current_user, RedirectResponse):
            return current_user
        remove_monitored_stock(resolved_settings.db_path, current_user["username"], symbol)
        return _redirect_with_message("/dashboard", "股票已删除")

    @app.get("/stocks/{symbol}", response_class=HTMLResponse)
    async def stock_detail(request: Request, symbol: str) -> Response:
        current_user = _require_login(request, resolved_settings)
        if isinstance(current_user, RedirectResponse):
            return current_user
        owner_username = current_user["username"]
        snapshot = get_snapshot(resolved_settings.db_path, owner_username, symbol)
        if snapshot is None:
            return _redirect_with_message("/dashboard", "当前账号下未找到该股票")
        snapshot = _decorate_snapshot_for_display(snapshot)
        monitor: StockMonitor = app.state.monitor
        chart_payload = await asyncio.to_thread(monitor.build_chart_payload, symbol)
        return templates.TemplateResponse(
            request,
            "stock_detail.html",
            _base_template_context(
                request,
                current_user,
                chart_payload=chart_payload,
                page_title=f"股票详情 {symbol}",
                snapshot=snapshot,
                symbol=symbol,
            ),
        )

    @app.post("/trades")
    async def create_trade_record(
        request: Request,
        symbol: str = Form(...),
        side: str = Form(...),
        price: float = Form(...),
        quantity: int = Form(...),
        note: str = Form(""),
    ) -> RedirectResponse:
        current_user = _require_login(request, resolved_settings)
        if isinstance(current_user, RedirectResponse):
            return current_user

        owner_username = current_user["username"]
        if side not in {"buy", "sell"}:
            return _redirect_with_message("/trades", "交易方向只能是买入或卖出")
        if not symbol.isdigit() or len(symbol) != 6:
            return _redirect_with_message("/trades", "请输入 6 位股票代码")
        price = round(float(price), 5)
        if price <= 0 or quantity <= 0:
            return _redirect_with_message("/trades", "价格和数量必须大于 0")

        try:
            add_monitored_stock(resolved_settings.db_path, owner_username, symbol)
            add_trade_record(
                resolved_settings.db_path,
                owner_username=owner_username,
                symbol=symbol,
                side=side,
                price=price,
                quantity=quantity,
                note=note.strip(),
            )
        except Exception:
            return _redirect_with_message(f"/trades?symbol={symbol}", "保存交易记录失败，请重试")
        return _redirect_with_message(f"/trades?symbol={symbol}", "交易记录已保存")

    @app.post("/trades/analyze")
    async def analyze_trade_plan(
        request: Request,
        symbol: str = Form(...),
    ) -> RedirectResponse:
        current_user = _require_login(request, resolved_settings)
        if isinstance(current_user, RedirectResponse):
            return current_user

        owner_username = current_user["username"]
        if not symbol.isdigit() or len(symbol) != 6:
            return _redirect_with_message("/trades", "请输入 6 位股票代码后再分析")

        trade_rows = list_trade_records_for_symbol(resolved_settings.db_path, owner_username, symbol)
        if not trade_rows:
            return _redirect_with_message(f"/trades?symbol={symbol}", "请先录入至少一条交易记录")

        monitor: StockMonitor = app.state.monitor
        trade_advisor: TradeAdvisor = app.state.trade_advisor

        try:
            snapshot = await asyncio.to_thread(monitor.refresh_symbol_snapshot, owner_username, symbol)
            snapshot_payload = {
                "symbol": snapshot.symbol,
                "display_name": snapshot.display_name,
                "latest_price": snapshot.latest_price,
                "ma_250": snapshot.ma_250,
                "ma_30w": snapshot.ma_30w,
                "ma_60w": snapshot.ma_60w,
                "boll_mid": snapshot.boll_mid,
                "boll_lower": snapshot.boll_lower,
                "boll_upper": snapshot.boll_upper,
                "dividend_yield": snapshot.dividend_yield,
                "quant_probability": snapshot.quant_probability,
                "quant_model_breakdown": snapshot.quant_model_breakdown,
                "trigger_state": snapshot.trigger_state,
                "trigger_detail": snapshot.trigger_detail,
                "updated_at": snapshot.updated_at.isoformat(),
            }
        except Exception:
            fallback_snapshot = get_snapshot(resolved_settings.db_path, owner_username, symbol)
            if fallback_snapshot is None or fallback_snapshot["latest_price"] is None:
                return _redirect_with_message(
                    f"/trades?symbol={symbol}",
                    "当前无法获取行情，请稍后再试",
                )
            snapshot_payload = dict(fallback_snapshot)
        result = await asyncio.to_thread(
            trade_advisor.analyze_symbol,
            symbol,
            snapshot_payload,
            trade_rows,
        )
        add_trade_analysis(
            resolved_settings.db_path,
            owner_username=owner_username,
            symbol=symbol,
            analysis_provider=result.provider,
            model_name=result.model_name,
            position_summary=build_position_summary(trade_rows),
            market_snapshot=snapshot_payload,
            analysis_json=result.analysis,
            status=result.status,
            error_message=result.error_message,
        )
        message = "交易复盘分析已更新"
        email_result = _send_trade_analysis_email_if_configured(
            db_path=resolved_settings.db_path,
            owner_username=owner_username,
            symbol=symbol,
        )
        if email_result is not None:
            message = f"{message}；{email_result}"
        if result.error_message:
            message = f"交易复盘分析已更新：{result.error_message}"
        return _redirect_with_message(f"/trades?symbol={symbol}", message)

    @app.post("/trades/email-analysis")
    async def email_trade_analysis(
        request: Request,
        symbol: str = Form(...),
    ) -> RedirectResponse:
        current_user = _require_login(request, resolved_settings)
        if isinstance(current_user, RedirectResponse):
            return current_user

        if not symbol.isdigit() or len(symbol) != 6:
            return _redirect_with_message("/trades", "请输入 6 位股票代码后再发送")

        result_message = _send_trade_analysis_email_if_configured(
            resolved_settings.db_path,
            current_user["username"],
            symbol,
        )
        if result_message is None:
            result_message = "请先生成至少一条复盘分析后再发送"
        return _redirect_with_message(f"/trades?symbol={symbol}", result_message)

    @app.post("/settings/email")
    async def update_email_settings(
        request: Request,
        recipient_email: str = Form(""),
        smtp_server: str = Form(""),
        sender_email: str = Form(""),
        sender_password: str = Form(""),
    ) -> RedirectResponse:
        current_user = _require_login(request, resolved_settings)
        if isinstance(current_user, RedirectResponse):
            return current_user

        save_email_settings(
            db_path=resolved_settings.db_path,
            owner_username=current_user["username"],
            recipient_email=recipient_email.strip(),
            smtp_server=smtp_server.strip(),
            sender_email=sender_email.strip(),
            sender_password=sender_password.strip(),
        )
        return _redirect_with_message("/settings", "邮箱配置已保存")

    @app.post("/settings/quant")
    async def update_quant_settings(
        request: Request,
        enabled: str | None = Form(None),
        probability_threshold: float = Form(90),
        selected_models: list[str] | None = Form(None),
        require_price_above_ma250: str | None = Form(None),
        require_weekly_bullish: str | None = Form(None),
        min_dividend_yield: float = Form(3.0),
        max_20d_volatility_pct: float = Form(4.0),
        min_20d_momentum_pct: float = Form(1.0),
        max_boll_deviation_pct: float = Form(4.0),
        support_zone_tolerance_pct: float = Form(3.0),
        min_reward_risk_ratio: float = Form(1.6),
        dcf_discount_rate_pct: float = Form(10.0),
        dcf_terminal_growth_pct: float = Form(3.0),
    ) -> RedirectResponse:
        current_user = _require_login(request, resolved_settings)
        if isinstance(current_user, RedirectResponse):
            return current_user

        normalized_models = list(normalize_selected_models(selected_models or []))
        threshold = max(50.0, min(99.0, float(probability_threshold)))
        strategy_params = normalize_strategy_params(
            {
                "require_price_above_ma250": require_price_above_ma250 is not None,
                "require_weekly_bullish": require_weekly_bullish is not None,
                "min_dividend_yield": min_dividend_yield,
                "max_20d_volatility": max_20d_volatility_pct / 100,
                "min_20d_momentum_pct": min_20d_momentum_pct / 100,
                "max_boll_deviation_pct": max_boll_deviation_pct / 100,
                "support_zone_tolerance_pct": support_zone_tolerance_pct / 100,
                "min_reward_risk_ratio": min_reward_risk_ratio,
                "dcf_discount_rate": dcf_discount_rate_pct / 100,
                "dcf_terminal_growth": dcf_terminal_growth_pct / 100,
            }
        )
        save_quant_settings(
            db_path=resolved_settings.db_path,
            owner_username=current_user["username"],
            enabled=enabled is not None,
            probability_threshold=threshold,
            selected_models=normalized_models,
            strategy_params=strategy_params,
        )
        snapshots = get_snapshots(resolved_settings.db_path, current_user["username"])
        for item in snapshots:
            try:
                await asyncio.to_thread(app.state.monitor.refresh_symbol_snapshot, current_user["username"], item["symbol"])
            except Exception:
                continue
        return _redirect_with_message("/settings", "量化提醒设置已保存")

    @app.post("/settings/admin-password")
    async def update_account_password_route(
        request: Request,
        current_password: str = Form(...),
        new_password: str = Form(...),
        confirm_password: str = Form(...),
    ) -> RedirectResponse:
        current_user = _require_login(request, resolved_settings)
        if isinstance(current_user, RedirectResponse):
            return current_user

        if not verify_password(current_password, current_user["password_hash"]):
            return _redirect_with_message("/settings", "当前密码不正确")
        if len(new_password) < 8:
            return _redirect_with_message("/settings", "新密码至少需要 8 位")
        if new_password != confirm_password:
            return _redirect_with_message("/settings", "两次输入的新密码不一致")
        if new_password == current_password:
            return _redirect_with_message("/settings", "新密码不能和当前密码相同")

        update_user_password(resolved_settings.db_path, current_user["username"], hash_password(new_password))
        refreshed_user = get_user(resolved_settings.db_path, current_user["username"])
        response = _redirect_with_message("/settings", "当前账号密码已更新，请牢记新密码")
        if refreshed_user is not None:
            response.set_cookie(
                resolved_settings.session_cookie_name,
                _build_session_cookie_value(refreshed_user),
                httponly=True,
                samesite="lax",
            )
        return response

    @app.post("/settings/users")
    async def create_user_route(
        request: Request,
        username: str = Form(...),
        password: str = Form(...),
        confirm_password: str = Form(...),
        is_admin: str | None = Form(None),
    ) -> RedirectResponse:
        current_user = _require_login(request, resolved_settings)
        if isinstance(current_user, RedirectResponse):
            return current_user
        if not current_user["is_admin"]:
            return _redirect_with_message("/settings", "只有管理员可以新增账号")

        normalized_username = username.strip()
        if len(normalized_username) < 3 or len(normalized_username) > 32:
            return _redirect_with_message("/settings", "账号长度需在 3 到 32 位之间")
        if not all(char.isalnum() or char in {"_", "-", "."} for char in normalized_username):
            return _redirect_with_message("/settings", "账号仅支持字母、数字、下划线、中划线和点")
        if len(password) < 8:
            return _redirect_with_message("/settings", "新账号密码至少需要 8 位")
        if password != confirm_password:
            return _redirect_with_message("/settings", "两次输入的账号密码不一致")
        if get_user(resolved_settings.db_path, normalized_username) is not None:
            return _redirect_with_message("/settings", "该账号已存在")
        try:
            create_user(
                resolved_settings.db_path,
                username=normalized_username,
                password_hash=hash_password(password),
                is_admin=is_admin is not None,
            )
        except sqlite3.IntegrityError:
            return _redirect_with_message("/settings", "该账号已存在")
        return _redirect_with_message("/settings", f"账号 {normalized_username} 已创建")

    @app.post("/settings/email/test")
    async def test_email(request: Request) -> RedirectResponse:
        current_user = _require_login(request, resolved_settings)
        if isinstance(current_user, RedirectResponse):
            return current_user
        email_settings = get_email_settings(resolved_settings.db_path, current_user["username"])
        result = send_message(
            email_settings,
            subject=f"[邮箱测试] {current_user['username']} SMTP 配置检查",
            body=build_test_email_body(current_user["username"], str(email_settings["recipient_email"] or "")),
        )
        if result.success:
            notice = f"测试邮件发送成功，请检查收件箱：{email_settings['recipient_email']}"
        else:
            notice = f"测试邮件发送失败：{result.error}"
        return _redirect_with_message("/settings", notice)

    @app.post("/alerts/{alert_id}/read")
    async def confirm_alert(request: Request, alert_id: int) -> RedirectResponse:
        current_user = _require_login(request, resolved_settings)
        if isinstance(current_user, RedirectResponse):
            return current_user
        mark_alert_as_read(resolved_settings.db_path, current_user["username"], alert_id)
        return _redirect_with_message("/dashboard", "提醒已确认")

    @app.get("/history", response_class=HTMLResponse)
    async def history_page(
        request: Request,
        symbol: str | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
        message: str | None = None,
        message_type: str = "info",
    ) -> Response:
        current_user = _require_login(request, resolved_settings)
        if isinstance(current_user, RedirectResponse):
            return current_user

        parsed_from = date.fromisoformat(date_from) if date_from else None
        parsed_to = date.fromisoformat(date_to) if date_to else None
        alerts = get_alert_history(
            resolved_settings.db_path,
            owner_username=current_user["username"],
            symbol=symbol or None,
            date_from=parsed_from,
            date_to=parsed_to,
        )
        return templates.TemplateResponse(
            request,
            "history.html",
            _base_template_context(
                request,
                current_user,
                alerts=alerts,
                message=message,
                message_type=message_type,
                date_from=date_from or "",
                date_to=date_to or "",
                page_title="历史提醒",
                symbol=symbol or "",
            ),
        )

    return app


async def _monitor_loop(app: FastAPI) -> None:
    settings: AppSettings = app.state.settings
    monitor: StockMonitor = app.state.monitor
    while True:
        await asyncio.to_thread(monitor.run_cycle)
        await asyncio.sleep(settings.refresh_interval_seconds)


def _build_system_status(monitor: StockMonitor, email_settings: object) -> dict[str, str]:
    mail_status = "已配置" if email_settings["smtp_server"] else "未配置"
    provider_status = "正常" if not monitor.last_error_message else f"异常：{monitor.last_error_message}"
    last_refresh = monitor.last_refresh_at.isoformat(sep=" ", timespec="seconds") if monitor.last_refresh_at else "尚未刷新"
    return {
        "server": "运行中",
        "provider": provider_status,
        "mail": mail_status,
        "last_refresh": last_refresh,
    }


def _base_template_context(request: Request, current_user: object, **kwargs: object) -> dict[str, object]:
    context = {
        "request": request,
        "current_user": current_user,
        "is_admin": bool(current_user["is_admin"]),
        "message_type": "info",
    }
    context.update(kwargs)
    return context


def _render_login_page(
    request: Request,
    error: str | None = None,
    message: str | None = None,
    unlock_username: str = "",
    status_code: int = 200,
) -> Response:
    return templates.TemplateResponse(
        request,
        "login.html",
        {
            "error": error,
            "message": message,
            "page_title": "登录",
            "request": request,
            "unlock_username": unlock_username,
        },
        status_code=status_code,
    )


def _resolve_unlock_email_target(email_settings: object) -> str:
    return str(email_settings["recipient_email"] or "").strip()


def _request_client_host(request: Request) -> str:
    forwarded_for = (request.headers.get("x-forwarded-for") or "").split(",")[0].strip()
    return forwarded_for or (request.client.host if request.client else "unknown")


def _request_user_agent(request: Request) -> str:
    return (request.headers.get("user-agent") or "unknown")[:120]


def _mask_email_address(email: str) -> str:
    if "@" not in email:
        return email
    local_part, domain = email.split("@", 1)
    if len(local_part) <= 2:
        masked_local = local_part[:1] + "*"
    else:
        masked_local = local_part[:2] + "***" + local_part[-1:]
    return f"{masked_local}@{domain}"


def _build_session_cookie_value(user: object) -> str:
    issued_at = int(datetime.now(UTC).timestamp())
    return f"{user['username']}|{str(user['password_hash'])[:24]}|{issued_at}"


def _get_authenticated_user(request: Request, settings: AppSettings) -> object | None:
    cookie_value = request.cookies.get(settings.session_cookie_name)
    request.state.auth_failure_reason = None
    if not cookie_value or "|" not in cookie_value:
        return None
    parts = cookie_value.split("|")
    if len(parts) < 2:
        return None
    username = parts[0]
    fingerprint = parts[1]
    issued_at_text = parts[2] if len(parts) >= 3 else None
    if not username:
        return None
    user = get_user(settings.db_path, username)
    if user is None:
        return None
    if not str(user["password_hash"]).startswith(fingerprint):
        return None
    if issued_at_text is not None:
        try:
            issued_at = int(issued_at_text)
        except ValueError:
            request.state.auth_failure_reason = "invalid"
            return None
        now_ts = int(datetime.now(UTC).timestamp())
        if now_ts - issued_at > SESSION_IDLE_TIMEOUT_SECONDS:
            request.state.auth_failure_reason = "idle_timeout"
            return None
    return user


def _serialize_latest_analysis_row(latest_analysis_row: object) -> dict[str, object]:
    market_snapshot = _decorate_snapshot_for_display(json.loads(latest_analysis_row["market_snapshot"]))
    analysis = json.loads(latest_analysis_row["analysis_json"])
    analysis.setdefault("buy_recommendation_level", market_snapshot.get("buy_recommendation_level", 5))
    analysis.setdefault("sell_recommendation_level", market_snapshot.get("sell_recommendation_level", 5))
    analysis.setdefault("buy_recommendation_level_label", market_snapshot.get("buy_recommendation_level_label", "中性"))
    analysis.setdefault("sell_recommendation_level_label", market_snapshot.get("sell_recommendation_level_label", "中性"))
    analysis.setdefault("recommendation_level_method", market_snapshot.get("recommendation_level_method", "基于多模型综合评分"))
    model_consensus = str(market_snapshot.get("model_consensus") or "多模型综合结论")
    watch_price_range = analysis.get('watch_price_range', '-')
    recommended_buy_price_range = analysis.get('recommended_buy_price_range', '-')
    recommended_sell_price_range = analysis.get('recommended_sell_price_range', '-')
    dcf_intrinsic_value = market_snapshot.get('dcf_intrinsic_value')
    dcf_valuation_gap_pct = market_snapshot.get('dcf_valuation_gap_pct')
    if dcf_intrinsic_value is not None:
        dcf_line = f"DCF：内在价值 {float(dcf_intrinsic_value):.2f}，偏差 {float(dcf_valuation_gap_pct or 0.0):.2f}%"
    else:
        dcf_line = f"DCF：{market_snapshot.get('dcf_reason') or '未启用 DCF 模型或当前估值数据不足'}"
    analysis["comprehensive_advice_text"] = (
        f"{model_consensus}；{analysis['position_advice']}；"
        f"推荐买入价 {recommended_buy_price_range}；"
        f"推荐卖出价 {recommended_sell_price_range}；"
        f"买入等级 {analysis['buy_recommendation_level']}/10；"
        f"卖出等级 {analysis['sell_recommendation_level']}/10；"
        f"观望关注价 {watch_price_range}。"
    )
    analysis["comprehensive_advice_card"] = {
        "conclusion": f"{model_consensus}；{analysis['position_advice']}",
        "buy": (
            f"买点：{recommended_buy_price_range}；"
            f"买入等级 {analysis['buy_recommendation_level']}/10（{analysis['buy_recommendation_level_label']}）；"
            f"关注位 {watch_price_range}"
        ),
        "sell": (
            f"卖点：{recommended_sell_price_range}；"
            f"卖出等级 {analysis['sell_recommendation_level']}/10（{analysis['sell_recommendation_level_label']}）"
        ),
        "dcf": dcf_line,
    }
    return {
        "symbol": latest_analysis_row["symbol"],
        "provider": latest_analysis_row["analysis_provider"],
        "model_name": latest_analysis_row["model_name"],
        "status": latest_analysis_row["status"],
        "error_message": latest_analysis_row["error_message"],
        "created_at": latest_analysis_row["created_at"],
        "market_snapshot": market_snapshot,
        "analysis": analysis,
    }


def _send_trade_analysis_email_if_configured(db_path: str, owner_username: str, symbol: str) -> str | None:
    latest_analysis_row = get_latest_trade_analysis(db_path, owner_username, symbol)
    if latest_analysis_row is None:
        return None

    latest_analysis = _serialize_latest_analysis_row(latest_analysis_row)
    latest_analysis["portfolio_profile"] = build_portfolio_profile(
        list_trade_records(db_path, owner_username, None),
        get_snapshots(db_path, owner_username),
    )
    email_settings = get_email_settings(db_path, owner_username)
    result = send_message(
        email_settings,
        subject=f"[交易复盘] {symbol} {latest_analysis['analysis']['judgment']}",
        body=build_trade_analysis_email_body(latest_analysis),
    )
    if result.success:
        return "复盘结果邮件已发送"
    return f"复盘结果邮件发送失败：{result.error}"


def _normalize_message_type(message_type: str | None) -> str:
    if message_type in {"success", "error", "warning", "info"}:
        return str(message_type)
    return "info"


def _guess_message_type(message: str) -> str:
    lowered = message.lower()
    error_keywords = ("失败", "错误", "不正确", "无效", "未找到", "不能", "不存在", "请先", "不一致", "已锁定")
    warning_keywords = ("部分", "稍后", "暂无", "暂未", "未配置", "待", "还没有")
    if any(keyword in message for keyword in error_keywords) or "error" in lowered or "failed" in lowered:
        return "error"
    if any(keyword in message for keyword in warning_keywords) or "warning" in lowered:
        return "warning"
    return "success"


def _require_login(request: Request, settings: AppSettings) -> object | RedirectResponse:
    current_user = _get_authenticated_user(request, settings)
    if current_user is not None:
        return current_user
    if getattr(request.state, "auth_failure_reason", None) == "idle_timeout":
        response = RedirectResponse(
            url="/login?message=%E7%94%B1%E4%BA%8E30%E5%88%86%E9%92%9F%E6%97%A0%E6%93%8D%E4%BD%9C%EF%BC%8C%E7%99%BB%E5%BD%95%E5%B7%B2%E8%BF%87%E6%9C%9F%EF%BC%8C%E8%AF%B7%E9%87%8D%E6%96%B0%E7%99%BB%E5%BD%95%E3%80%82",
            status_code=303,
        )
        response.delete_cookie(settings.session_cookie_name)
        return response
    return RedirectResponse(url="/login", status_code=303)


def _redirect_with_message(path: str, message: str, message_type: str | None = None) -> RedirectResponse:
    resolved_type = _normalize_message_type(message_type or _guess_message_type(message))
    separator = "&" if "?" in path else "?"
    return RedirectResponse(
        url=f"{path}{separator}message={quote(message)}&message_type={quote(resolved_type)}",
        status_code=303,
    )


def _login_subject(request: Request, username: str = "") -> str:
    client_host = _request_client_host(request)
    user_agent = _request_user_agent(request)
    return f"{username.lower()}|{client_host}|{user_agent}"


async def _load_dashboard_snapshots(app: FastAPI, owner_username: str) -> list[object]:
    settings: AppSettings = app.state.settings
    monitor: StockMonitor = app.state.monitor
    snapshots = get_snapshots(settings.db_path, owner_username)
    refresh_symbols = [
        item["symbol"]
        for item in snapshots
        if item["latest_price"] is None or not float(item["boll_upper"] or 0)
    ]
    if not refresh_symbols:
        return snapshots

    for symbol in refresh_symbols:
        try:
            await asyncio.to_thread(monitor.refresh_symbol_snapshot, owner_username, symbol)
        except Exception:
            continue
    return get_snapshots(settings.db_path, owner_username)


def run() -> None:
    settings = load_settings()
    uvicorn.run(create_app(settings), host=settings.host, port=settings.port)
